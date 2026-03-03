"""
RBI Handbook of Statistics — XLSX downloader and parser.

Downloads Excel tables from:
  1. Handbook of Statistics on Indian Economy
     https://www.rbi.org.in/Scripts/AnnualPublications.aspx?head=Handbook+of+Statistics+on+Indian+Economy
  2. Handbook of Statistics on Indian States
     https://www.rbi.org.in/Scripts/AnnualPublications.aspx?head=Handbook+of+Statistics+on+Indian+States

The RBI website uses Imperva anti-bot protection. This module uses a
requests.Session() to first visit the publication page (collecting cookies),
then download individual XLSX table files.

Table URLs contain unpredictable hashes that change with each edition,
so we scrape the HTML page to discover current download links.
"""

import io
import logging
import re
import time
from typing import Any

import openpyxl
import requests
from bs4 import BeautifulSoup

logger = logging.getLogger(__name__)

_ECONOMY_URL = (
    "https://www.rbi.org.in/Scripts/AnnualPublications.aspx"
    "?head=Handbook+of+Statistics+on+Indian+Economy"
)
_STATES_URL = (
    "https://www.rbi.org.in/Scripts/AnnualPublications.aspx"
    "?head=Handbook+of+Statistics+on+Indian+States"
)

_USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
)

# Delay between downloads to avoid triggering rate limits
_DOWNLOAD_DELAY = 1.5  # seconds


def _create_session() -> requests.Session:
    """Create a session with browser-like headers."""
    s = requests.Session()
    s.headers.update({
        "User-Agent": _USER_AGENT,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
    })
    return s


def _scrape_table_urls(session: requests.Session, page_url: str) -> dict[int, str]:
    """
    Scrape a Handbook publication page for XLSX download URLs.

    Returns dict mapping table number → full download URL.
    Only captures the first occurrence of each table number
    (some tables appear multiple times as 19T, 119T, 219T etc.)
    """
    r = session.get(page_url, timeout=30)
    r.raise_for_status()

    soup = BeautifulSoup(r.text, "html.parser")
    links = soup.find_all("a", href=True)

    table_urls: dict[int, str] = {}
    for link in links:
        href = link["href"]
        if ".XLSX" not in href.upper():
            continue

        # Extract table number from filename like "43T_29082025...XLSX"
        filename = href.rsplit("/", 1)[-1]
        match = re.match(r"^(\d+)T_", filename)
        if not match:
            continue

        table_num = int(match.group(1))
        if table_num not in table_urls:
            table_urls[table_num] = href

    logger.info(f"  Handbook: found {len(table_urls)} table URLs from {page_url.split('head=')[1]}")
    return table_urls


def download_table(
    session: requests.Session,
    url: str,
    table_num: int,
) -> openpyxl.Workbook | None:
    """
    Download an XLSX table and return an openpyxl Workbook.

    Returns None if download fails or file is not a valid XLSX.
    """
    try:
        r = session.get(url, timeout=30)
        if r.status_code != 200:
            logger.warning(f"  Table {table_num}: HTTP {r.status_code}")
            return None

        # Verify it's actually an XLSX (PK zip header), not an HTML error page
        if r.content[:2] != b"PK":
            logger.warning(f"  Table {table_num}: not a valid XLSX (got HTML/other)")
            return None

        wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
        return wb

    except Exception as e:
        logger.warning(f"  Table {table_num}: download failed — {e}")
        return None


def fetch_economy_tables(table_numbers: list[int]) -> dict[int, openpyxl.Workbook]:
    """
    Download specific tables from the Indian Economy Handbook.

    Returns dict mapping table number → Workbook for successfully downloaded tables.
    """
    session = _create_session()
    table_urls = _scrape_table_urls(session, _ECONOMY_URL)

    results: dict[int, openpyxl.Workbook] = {}
    for num in table_numbers:
        url = table_urls.get(num)
        if not url:
            logger.warning(f"  Table {num}: not found in Handbook page")
            continue

        time.sleep(_DOWNLOAD_DELAY)
        wb = download_table(session, url, num)
        if wb:
            results[num] = wb
            logger.info(f"  Table {num}: downloaded ({', '.join(wb.sheetnames)})")

    return results


def fetch_states_tables(table_numbers: list[int]) -> dict[int, openpyxl.Workbook]:
    """
    Download specific tables from the Indian States Handbook.

    Returns dict mapping table number → Workbook for successfully downloaded tables.
    """
    session = _create_session()
    table_urls = _scrape_table_urls(session, _STATES_URL)

    results: dict[int, openpyxl.Workbook] = {}
    for num in table_numbers:
        url = table_urls.get(num)
        if not url:
            logger.warning(f"  Table {num}: not found in States Handbook page")
            continue

        time.sleep(_DOWNLOAD_DELAY)
        wb = download_table(session, url, num)
        if wb:
            results[num] = wb
            logger.info(f"  Table {num}: downloaded ({', '.join(wb.sheetnames)})")

    return results


def parse_state_table(
    wb: openpyxl.Workbook,
    value_columns: list[int] | None = None,
) -> list[dict[str, Any]]:
    """
    Parse a state-level Handbook table (rows = states, columns = years).

    These tables have a consistent format:
      Row 2: Title
      Row 3: Unit
      Row 4: Header row with "State/Union Territory" + base year label
      Row 5: Year labels (e.g., "2011-12", "2012-13", ...)
      Row 6+: Data rows (state name in col B, values in cols C+)

    Returns list of {state: str, years: {year: value, ...}} dicts.

    Combines data from all sheets (tables split across sheets like T_21(i), T_21(ii)).
    """
    state_data: dict[str, dict[str, float]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Find the year header row (contains fiscal year strings like "2011-12")
        year_labels: dict[int, str] = {}
        for row in ws.iter_rows(min_row=3, max_row=8, values_only=False):
            for cell in row:
                val = cell.value
                if isinstance(val, str) and re.match(r"\d{4}-\d{2}", val):
                    year_labels[cell.column] = val

            if year_labels:
                break

        if not year_labels:
            continue

        # Find first data row: a row after the year headers that has a state
        # name in column B and numeric data in year columns.
        # Year labels are typically in rows 4-5; data starts at row 6+.
        year_header_row = min(
            cell.row
            for row in ws.iter_rows(min_row=3, max_row=8, values_only=False)
            for cell in row
            if isinstance(cell.value, str) and re.match(r"\d{4}-\d{2}", cell.value)
        )
        data_start_row = year_header_row + 1  # default: row after year labels

        for row_idx in range(year_header_row + 1, ws.max_row + 1):
            cell_b = ws.cell(row=row_idx, column=2).value
            if not isinstance(cell_b, str) or not cell_b.strip():
                continue
            stripped = cell_b.strip()
            # Skip header/label rows
            if stripped.isdigit() or stripped.startswith(("TABLE", "(", "Base")):
                continue
            if stripped in ("State/Union Territory",):
                continue
            # Check if row has numeric data in year columns
            has_data = any(
                ws.cell(row=row_idx, column=c).value is not None
                for c in year_labels
            )
            if has_data:
                data_start_row = row_idx
                break

        for row in ws.iter_rows(min_row=data_start_row, max_row=ws.max_row, values_only=False):
            state_name = row[1].value if len(row) > 1 else None  # Column B (index 1)
            if not isinstance(state_name, str) or not state_name.strip():
                continue

            state_name = state_name.strip()

            # Skip footnote/source rows
            if state_name.startswith(("-", "Note", "Source", "*", "#", "Also")):
                continue
            # Skip numeric-only rows (column number headers like "1")
            if state_name.isdigit():
                continue

            years = state_data.setdefault(state_name, {})

            for col, year_label in year_labels.items():
                cell = ws.cell(row=row[0].row, column=col)
                val = cell.value
                if val is None or val == "-" or val == "--" or val == "..":
                    continue
                try:
                    years[year_label] = float(val)
                except (ValueError, TypeError):
                    pass

    return [
        {"state": state, "years": years}
        for state, years in state_data.items()
        if years  # Skip states with no data
    ]


def parse_rate_table(wb: openpyxl.Workbook) -> list[dict[str, Any]]:
    """
    Parse Table 43 (monetary policy rates) or similar event-driven tables.

    These tables have dates in column B and rates that change event-by-event.
    '-' means "no change" for that field.

    Combines all sheets and returns chronologically sorted list of:
    {date, bankRate, repo, reverseRepo, sdf, msf, crr, slr}
    """
    events: list[dict[str, Any]] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        for row in ws.iter_rows(min_row=7, max_row=ws.max_row, values_only=False):
            date_cell = row[1].value if len(row) > 1 else None  # Column B

            # Skip non-date rows
            if date_cell is None:
                continue
            if isinstance(date_cell, str):
                if date_cell.startswith(("Note", "Source", "(", "Also")):
                    break
                continue

            # Should be a datetime
            import datetime
            if not isinstance(date_cell, datetime.datetime):
                continue

            date_str = date_cell.strftime("%Y-%m-%d")

            def _get_val(col_idx: int) -> float | None:
                if col_idx >= len(row):
                    return None
                v = row[col_idx].value
                if v is None or v == "-" or v == "--":
                    return None
                try:
                    return float(v)
                except (ValueError, TypeError):
                    return None

            events.append({
                "date": date_str,
                "bankRate": _get_val(2),     # Column C
                "repo": _get_val(3),          # Column D
                "reverseRepo": _get_val(4),   # Column E
                "sdf": _get_val(5),           # Column F
                "msf": _get_val(6),           # Column G
                "crr": _get_val(7),           # Column H
                "slr": _get_val(8),           # Column I
            })

    events.sort(key=lambda e: e["date"])
    return events


def parse_interest_rate_table(wb: openpyxl.Workbook) -> list[dict[str, Any]]:
    """
    Parse Table 62 (structure of interest rates).

    Returns list of {year, callMoneyRate, depositRate1to3, lendingRate, walrOutstanding, mclr1Year}
    """
    ws = wb.active
    results: list[dict[str, Any]] = []

    for row in ws.iter_rows(min_row=9, max_row=ws.max_row, values_only=False):
        year_cell = row[1].value if len(row) > 1 else None
        if not isinstance(year_cell, str) or not re.match(r"\d{4}-\d{2}", year_cell):
            continue

        def _get_numeric(col_idx: int) -> float | None:
            if col_idx >= len(row):
                return None
            v = row[col_idx].value
            if v is None or v == "--" or v == "-":
                return None
            if isinstance(v, (int, float)):
                return float(v)
            # Handle ranges like "8.00-8.75" — take midpoint
            if isinstance(v, str) and "-" in v:
                parts = v.split("-")
                try:
                    vals = [float(p) for p in parts]
                    return round(sum(vals) / len(vals), 2)
                except ValueError:
                    return None
            try:
                return float(v)
            except (ValueError, TypeError):
                return None

        results.append({
            "year": year_cell,
            "callMoneyRate": _get_numeric(2),       # Col C
            "depositRate1to3": _get_numeric(4),       # Col E (1-3yr term deposits)
            "lendingRate": _get_numeric(7),           # Col H
            "mclr1Year": _get_numeric(8),             # Col I
            "walrOutstanding": _get_numeric(9),       # Col J
            "walrFresh": _get_numeric(10),            # Col K
            "wadtdr": _get_numeric(12),               # Col M
        })

    return results


def parse_forex_table(wb: openpyxl.Workbook) -> list[dict[str, Any]]:
    """
    Parse Table 147 (foreign exchange reserves).

    Returns list of {year, totalCrore, totalUsdMillion}
    """
    ws = wb.active
    results: list[dict[str, Any]] = []

    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, values_only=False):
        year_cell = row[1].value if len(row) > 1 else None
        if not isinstance(year_cell, str) or not re.match(r"\d{4}-\d{2}", year_cell):
            continue

        def _get_val(col_idx: int) -> float | None:
            if col_idx >= len(row):
                return None
            v = row[col_idx].value
            if v is None or v == "-" or v == "--":
                return None
            try:
                return float(v)
            except (ValueError, TypeError):
                return None

        results.append({
            "year": year_cell,
            "totalCrore": _get_val(10),       # Col K (Total ₹ Crore)
            "totalUsdMillion": _get_val(11),   # Col L (Total US$ Million)
        })

    return results
